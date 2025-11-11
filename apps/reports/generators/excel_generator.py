"""
Generador de Reportes en Excel

Utiliza openpyxl para generar archivos Excel (.xlsx) con formato.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from .base import BaseReportGenerator
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)


class ExcelReportGenerator(BaseReportGenerator):
    """Generador de reportes en Excel usando openpyxl"""

    def __init__(self, title="Reporte", format_type="excel", user=None):
        super().__init__(title, format_type)
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remover hoja por defecto
        self.current_sheet = None
        self.user = user

    def create_sheet(self, sheet_name: str):
        """
        Crear una nueva hoja en el workbook

        Args:
            sheet_name: Nombre de la hoja

        Returns:
            Worksheet: Hoja creada
        """
        self.current_sheet = self.workbook.create_sheet(title=sheet_name)
        return self.current_sheet

    def add_title_row(self, sheet, title: str, row: int = 1):
        """
        Agregar fila de título con formato

        Args:
            sheet: Hoja de Excel
            title: Título a agregar
            row: Número de fila
        """
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Merge cells para el título (asumiendo 5 columnas)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    def add_section(self, title: str, content: str) -> None:
        """
        Implementar método abstracto: Agregar una sección al reporte.

        Args:
            title: Título de la sección
            content: Contenido de la sección
        """
        if not self.current_sheet:
            self.create_sheet("Reporte")

        # Encontrar la siguiente fila vacía
        next_row = self.current_sheet.max_row + 1

        # Agregar título de sección
        title_cell = self.current_sheet.cell(row=next_row, column=1, value=title)
        title_cell.font = Font(size=12, bold=True)
        next_row += 1

        # Agregar contenido
        content_cell = self.current_sheet.cell(row=next_row, column=1, value=content)
        content_cell.alignment = Alignment(wrap_text=True)

    def add_table(self, data: List[Dict[str, Any]], headers: Optional[List[str]] = None, start_row: int = 1) -> None:
        """
        Implementar método abstracto: Agregar una tabla con datos y enumeración.

        Args:
            data: Lista de diccionarios con los datos
            headers: Lista de encabezados (opcional)
            start_row: Fila donde empezar la tabla
        """
        if not self.current_sheet:
            self.create_sheet("Datos")

        if not data:
            return

        # Usar headers si están provided, si no extraer del primer diccionario
        if headers is None:
            headers = list(data[0].keys()) if data else []

        # Agregar columna de enumeración
        headers_with_enum = ['#'] + headers

        # Estilos para encabezados con colores rose/cream
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="CFA195", end_color="CFA195", fill_type="solid")  # color-accent-rose
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin', color='A59383'),
            right=Side(style='thin', color='A59383'),
            top=Side(style='thin', color='A59383'),
            bottom=Side(style='thin', color='A59383')
        )

        # Escribir encabezados
        for col_idx, header in enumerate(headers_with_enum, start=1):
            cell = self.current_sheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Escribir datos con enumeración
        data_font = Font(name='Arial', size=9)
        data_alignment = Alignment(horizontal="left", vertical="center")
        enum_alignment = Alignment(horizontal="center", vertical="center")
        
        for row_idx, row_data in enumerate(data, start=start_row + 1):
            # Columna de enumeración
            enum_cell = self.current_sheet.cell(row=row_idx, column=1, value=row_idx - start_row)
            enum_cell.font = data_font
            enum_cell.alignment = enum_alignment
            enum_cell.border = border
            
            # Datos
            for col_idx, header in enumerate(headers, start=2):
                value = row_data.get(header, '')
                cell = self.current_sheet.cell(row=row_idx, column=col_idx, value=str(value))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border

                # Alternar colores: blanco y cream (#E2B8AD)
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E2B8AD", end_color="E2B8AD", fill_type="solid")
                    enum_cell.fill = PatternFill(start_color="E2B8AD", end_color="E2B8AD", fill_type="solid")

        # Ajustar ancho de columnas
        self.current_sheet.column_dimensions[get_column_letter(1)].width = 6  # Columna #
        
        for col_idx, header in enumerate(headers, start=2):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            for row_data in data:
                val_length = len(str(row_data.get(header, '')))
                if val_length > max_length:
                    max_length = val_length
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            self.current_sheet.column_dimensions[column_letter].width = adjusted_width

    def generate(self) -> bytes:
        """
        Implementar método abstracto: Generar el Excel y retornar bytes.

        Returns:
            bytes: Contenido del archivo Excel
        """
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def save(self, filename: str):
        """Guardar el Excel en un archivo"""
        self.workbook.save(filename)
        return filename
